#!/usr/bin/python
# -*- coding: utf-8 -*-

# Yemekbot 
# - Yekta Leblebici <yekta@iamyekta.com>

### Imports
import time
from slackclient import SlackClient
from openpyxl import load_workbook
from datetime import datetime
from ast import literal_eval

### Slack initialization
SECRET_KEY = "topsecret"
sc = SlackClient(SECRET_KEY)
BAD_WORDS = ['SPAGETTİ', 'FASULYE', 'BROKOLİ', 'SEBZE', 'PATLICAN', 'KEREVİZ', 'PIRASA', 'MAKARNA', 'ISPANAK', 'KARNABAHAR', 'TÜRLÜ']
GOOD_WORDS = ['SOTE', 'KÖFTE', 'KIZARTMA', 'PİLAV', 'PİRZOLA', 'BALIK', 'DÖNER', 'TAVUK', 'PEYNİRLİ', 'BÖREĞİ', 'BÖREK']
CHANNEL = "#random"

### Functions
def get_menu(fname="yemek.xlsx", date=None):
    """
    Parses an XLSX file and gets what's on the menu for the day. If date not set, defaults to today.
    """
    if date is None:
        date = datetime.strftime(datetime.today(), "%d-%m-%Y")

    try:
        date = datetime.strptime(date, "%d-%m-%Y")
    except:
        return False
    
    wb = load_workbook('yemek.xlsx')
    sheets = []
    outList = []

    for sheet in wb.get_sheet_names():
        ws = wb.get_sheet_by_name(sheet)
        for row in ws.iter_rows(min_row=1, max_col=1, max_row=50):
            for cell in row:
                if (cell.value == date):
                    for cell in ws['B' + str(cell.row):'F' + str(cell.row)][0]:
                        if cell.value is not None:
                            outList.append(cell.value.encode('utf-8'))
                    if outList != []:
                        return outList
    return False


def listen_channel(channel=CHANNEL):
    """
    Makes RTM connection to Slack and listens to commands. 

    """
    if sc.rtm_connect():
        while True:
            try:
                data = literal_eval(str(sc.rtm_read()))
                if len(data) > 0:
                    if 'text' in data[0]:
                            message =(data[0]['text'])
                            if (message[:6] == "!yemek"):
                                if (message == "!yemek"):
                                    post_slack(get_menu())
                                else:
                                    post_slack(get_menu(date=message[7:]))
            except:
                print("Connection interrupted.")
		sc.rtm_connect()
            time.sleep(1)
    else:
        print("Connection failed.")

def eval_food(item):
    """
    Evaluates food on a purely objective and scientific manner. Returns score as an Integer.

    BAD_WORDS take precedence over GOOD_WORDS on a single item.
    """
    for word in BAD_WORDS:
        if word in item:
            return -1
    
    for word in GOOD_WORDS:
        if word in item:
            return 2
    return 0

def eval_menu(foodScore):
    """
    Evaluates menu from its aggregated score and returns "chef's recommendation" as a String.
    
    """
    if (foodScore >= 0):
        return "Şefin tavsiyesi: Yemekhane. Skor: " + str(foodScore)
    else:
        return "Şefin tavsiyesi: Kentpark. Skor: " + str(foodScore)

#print(sc.api_call("users.list"))

def post_slack(foodList, channel=CHANNEL):
    """
    Generates a full string from menu items (foodList) and eval_food() result and posts it on Slack channel.

    """
    outBuf = ""
    foodScore = 0

    if (foodList == False):
        sc.api_call("chat.postMessage",
        channel=channel,
        as_user="true",
        text="Geçersiz tarih veya bu tarih için menü bilgisi bulunamadı.")
        return False

    for item in foodList:
        outBuf += item + ", "
        foodScore = foodScore + eval_food(item)
    

    sc.api_call("chat.postMessage",
        channel=channel,
        as_user="true",
        text=outBuf)

    sc.api_call("chat.postMessage",
        channel=channel,
        as_user="true",
        text=eval_menu(foodScore))

    return True

### Execution
print get_menu()
listen_channel()
#post_slack(get_menu(), channel="U3WAL8ZML")

#TODO testing?
#TODO config management?
